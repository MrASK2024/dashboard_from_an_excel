import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import time
import plotly.express as px
from datetime import datetime, timedelta
from config import EXCEL_FILE_PATH

st.set_page_config(layout="wide")
st.markdown(f"<h1 style='text-align: center;'>Статистика по группам</h1>", unsafe_allow_html=True)

if 'previous_data' not in st.session_state:
    st.session_state.previous_data = {}

if 'changed_cells' not in st.session_state:
    st.session_state.changed_cells = {}

if 'change_history' not in st.session_state:
    st.session_state.change_history = {}

# Настройки
EXCEL_FILE_PATH = EXCEL_FILE_PATH
REFRESH_INTERVAL = 10  # Интервал обновления страницы в секундах
TABLE_COLUMNS_PER_ROW = 4  # Количество таблиц в строке
CHART_WIDTH = 600
CHART_HEIGHT = 600

def load_excel_data(url):
    """Загружает данные из Excel файла."""
    wb = load_workbook(url, data_only=True)
    ws = wb.active
    return ws

def parse_groups(ws):
    """Парсит данные из Excel и возвращает словарь с группами."""
    groups = {}
    current_group = None

    for row in ws.iter_rows(min_row=2):
        row_values = [cell.value for cell in row]
        if row_values[0] and "итого за группировку" in str(row_values[0]).lower():
            current_group = row_values[0].split()[3].strip('"')
            groups[current_group] = {
                "header": ["Полученных", "", "", "Отремонтированных НСУ", "Отремонтировано DJI", "Сданные видео", "Сданные стикеры"],
                "subheader": ["", "День", "Ночь", "", "", "", ""],
                "data": []
            }
        elif row_values[0] and "итого" in str(row_values[0]).lower():
            if current_group:
                groups[current_group]["data"].append(["ИТОГО"] + row_values[2:])
        elif current_group:
            groups[current_group]["data"].append(row_values[1:])

    return groups

def track_changes(groups):
    """Отслеживает изменения в данных и сохраняет их в session_state."""
    current_changed_cells = {}
    current_time = datetime.now()

    # Очистка устаревших данных (старше суток)
    for key in list(st.session_state.change_history.keys()):
        changes = st.session_state.change_history[key]
        changes = [change for change in changes if current_time - datetime.strptime(change["time"], "%Y-%m-%d %H:%M:%S") < timedelta(days=1)]
        if changes:
            st.session_state.change_history[key] = changes
        else:
            del st.session_state.change_history[key]

    for group_name, group_data in groups.items():
        if group_name in st.session_state.previous_data:
            previous_group_data = st.session_state.previous_data[group_name]
            for i, row in enumerate(group_data["data"]):
                for j, cell_value in enumerate(row):
                    if i < len(previous_group_data["data"]) and j < len(previous_group_data["data"][i]):
                        if cell_value != previous_group_data["data"][i][j]:
                            current_changed_cells[(group_name, i, j)] = True
                            if (group_name, i, j) not in st.session_state.change_history:
                                st.session_state.change_history[(group_name, i, j)] = []
                            st.session_state.change_history[(group_name, i, j)].append({
                                "time": current_time.strftime("%Y-%m-%d %H:%M:%S"),
                                "value": cell_value
                            })

    st.session_state.previous_data = groups
    st.session_state.changed_cells = current_changed_cells

def render_group_table(group_name, group_data):
    """Отображает таблицу для одной группы."""
    html_table = """
    <style>
        th, td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: center;
            font-size: 16px;
        }
        th {
            background-color: #f2f2f2;
            text-align: center;
            font-size: 16px;
        }
        table {
            width: 100%;
            margin-bottom: 20px;
        }
        .changed {
            background-color: #1e8c45;
            animation: blink 1s ease-in-out;
        }
        @keyframes blink {
            0% { background-color: #1e8c45; }
            50% { background-color: #ffffff; }
            100% { background-color: #1e8c45; }
        }
    </style>
    <table>
    """
    html_table += "<tr>"
    html_table += "<th colspan='3' style='text-align: center;'>Полученных</th>"
    for header in group_data["header"][3:]:
        html_table += f"<th>{header}</th>"
    html_table += "</tr>"
    html_table += "<tr>"
    for subheader in group_data["subheader"]:
        html_table += f"<th>{subheader}</th>"
    html_table += "</tr>"
    for i, row in enumerate(group_data["data"]):
        html_table += "<tr>"
        for j, cell_value in enumerate(row):
            if (group_name, i, j) in st.session_state.changed_cells:
                html_table += f"<td class='changed'>{cell_value if cell_value is not None else ''}</td>"
            else:
                html_table += f"<td>{cell_value if cell_value is not None else ''}</td>"
        html_table += "</tr>"
    html_table += "</table>"
    st.markdown(html_table, unsafe_allow_html=True)

def render_pie_chart(group_name, group_data):
    """Отображает круговую диаграмму для одной группы."""
    all_data = []
    for row in group_data["data"]:
        if row[0] == "ИТОГО":
            row_values = [0 if value is None else value for value in row[1:]]
            total = sum(row_values)
            if total == 0:
                continue
            all_data.append({
                "Группа": group_name,
                "Полученных (День)": (row_values[0] / total * 100) if row_values[0] is not None else 0,
                "Полученных (Ночь)": (row_values[1] / total * 100) if row_values[1] is not None else 0,
                "Отремонтированных НСУ": (row_values[2] / total * 100) if row_values[2] is not None else 0,
                "Отремонтировано DJI": (row_values[3] / total * 100) if row_values[3] is not None else 0,
                "Сданные видео": (row_values[4] / total * 100) if row_values[4] is not None else 0,
                "Сданные стикеры": (row_values[5] / total * 100) if row_values[5] is not None else 0
            })

    if all_data:
        df = pd.DataFrame(all_data)
        df_melted = df.melt(id_vars=["Группа"], var_name="Категория", value_name="Доля (%)")
        fig = px.pie(df_melted, values="Доля (%)", names="Категория", 
                     title=f"Доли данных для {group_name} (в %)",
                     hole=0.3, width=800, height=800)
        fig.update_layout(
            title={
                'text': f"Доли данных для {group_name} (в %)",
                'y': 0.95,
                'x': 0.5,
                'xanchor': 'center',
                'yanchor': 'top',
                'font': {
                    'size': 24,
                    'family': "Arial",
                    'color': "black"
                }
            },
            legend={
                'font': {
                    'size': 24,
                    'family': "Arial",
                    'color': "black"
                }
            }
        )
        st.plotly_chart(fig, use_container_width=True)

def render_groups(groups):
    """Отображает таблицы и диаграммы для всех групп."""
    num_groups = len(groups)
    rows = (num_groups + TABLE_COLUMNS_PER_ROW - 1) // TABLE_COLUMNS_PER_ROW

    for row_index in range(rows):
        cols = st.columns(TABLE_COLUMNS_PER_ROW)
        for col_index in range(TABLE_COLUMNS_PER_ROW):
            group_index = row_index * TABLE_COLUMNS_PER_ROW + col_index
            if group_index < num_groups:
                group_name = list(groups.keys())[group_index]
                group_data = groups[group_name]
                with cols[col_index]:
                    st.markdown(f"<h4 style='text-align: center; font-size: 30px;'>{group_name}</h4>", unsafe_allow_html=True)
                    render_group_table(group_name, group_data)
                    render_pie_chart(group_name, group_data)

def render_daily_dynamics(groups):
    """Отображает динамику за день."""
    st.markdown("<h1 style='text-align: center;'>Динамика за день</h1>", unsafe_allow_html=True)

    if 'initial_values' not in st.session_state:
        st.session_state.initial_values = {}

    diff_data = []
    for group_name, group_data in groups.items():
        for row in group_data["data"]:
            if row[0] == "ИТОГО":
                current_values = row[1:]
                if group_name not in st.session_state.initial_values:
                    st.session_state.initial_values[group_name] = current_values
                initial_values = st.session_state.initial_values[group_name]
                diff_values = [current - initial if current is not None and initial is not None else 0 
                              for current, initial in zip(current_values, initial_values)]
                diff_data.append({
                    "Группа": group_name,
                    "Полученных (День)": diff_values[0],
                    "Полученных (Ночь)": diff_values[1],
                    "Отремонтированных НСУ": diff_values[2],
                    "Отремонтировано DJI": diff_values[3],
                    "Сданные видео": diff_values[4],
                    "Сданные стикеры": diff_values[5]
                })
                break

    if diff_data:
        df_diff = pd.DataFrame(diff_data)
        col1, col2 = st.columns([1, 2])

        with col1:
            st.markdown("<h4 style='text-align: center;'>Статистика за день</h4>", unsafe_allow_html=True)
            html_table = """
            <style>
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: center;
                    font-size: 30px;
                }
                th {
                    background-color: #f2f2f2;
                    text-align: center;
                    font-size: 30px;
                    vertical-align: middle;
                }
                table {
                    width: 90%;
                    margin-left: auto;
                    margin-right: auto;
                    margin-bottom: 30px;
                }
                .changed {
                    background-color: #1e8c45;
                    animation: blink 1s ease-in-out;
                }
                @keyframes blink {
                    0% { background-color: #1e8c45; }
                    50% { background-color: #ffffff; }
                    100% { background-color: #1e8c45; }
                }
            </style>
            <table>
            """
            html_table += "<tr>"
            for col in df_diff.columns:
                html_table += f"<th>{col}</th>"
            html_table += "</tr>"
            for _, row in df_diff.iterrows():
                html_table += "<tr>"
                for value in row:
                    html_table += f"<td>{value}</td>"
                html_table += "</tr>"
            html_table += "</table>"
            st.markdown(html_table, unsafe_allow_html=True)

        with col2:
            st.markdown("<h4 style='text-align: center;'>Динамика</h4>", unsafe_allow_html=True)
            df_melted = df_diff.melt(id_vars=["Группа"], var_name="Категория", value_name="Разница")
            fig = px.bar(df_melted, x="Группа", y="Разница", color="Категория",
                         labels={"Группа": " ", "Разница": "Разница", "Категория": " "},
                         barmode="group", width=300, height=CHART_HEIGHT)
            
            unique_groups = df_melted["Группа"].unique()
            for i, group in enumerate(unique_groups):
                if i > 0:  # Не добавляем линию перед первой группой
                    # Позиция линии между группами
                    line_position = i - 0.5
                    fig.add_vline(x=line_position, line_width=2, line_dash="dash", line_color="gray")
            
            fig.update_layout(
                legend={
                    'font': {
                        'size': 24,
                        'family': "Arial",
                        'color': "black"
                    }
                },
                xaxis={
                    'tickfont': {  # Стиль цифр на оси X
                        'size': 24,
                        'family': "Arial",
                        'color': "black"
                    }
                },
                yaxis={
                    'title': {
                        'text': 'Разница',  # Подпись оси Y
                        'font': {
                            'size': 24,
                            'family': "Arial",
                            'color': "black"
                        }
                    },
                    'tickfont': {  # Стиль цифр на оси Y
                        'size': 24,
                        'family': "Arial",
                        'color': "black"
                    }
                }
            )
            st.plotly_chart(fig, use_container_width=True)

        st.markdown("<div style='margin-bottom: 100px;'></div>", unsafe_allow_html=True)

def main():
    """Основная функция для запуска приложения."""
    ws = load_excel_data(EXCEL_FILE_PATH)
    groups = parse_groups(ws)
    if not groups:
        st.error("Нет данных для отображения.")
        return

    track_changes(groups)
    render_groups(groups)
    render_daily_dynamics(groups)

    time.sleep(REFRESH_INTERVAL)
    st.rerun()

if __name__ == "__main__":
    main()